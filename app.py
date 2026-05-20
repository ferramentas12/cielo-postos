from flask import Flask, request, send_file, jsonify
from flask_cors import CORS
import openpyxl
import io
import json
import re
from datetime import date, timedelta

app = Flask(__name__, static_folder='.', static_url_path='')
CORS(app)

POSTO_CONFIGS = {
    'VILAS': {
        'col_map': {'AMEX':2,'HIPER':3,'MASTER CREDITO':4,'MASTER DEBITO':5,'VISA CREDITO':6,'ELO CREDITO':7,'VISA DEBITO':8,'ELO DEBITO':9,'PIX':11},
        'taxas': {2:0.0210,3:0.0249,4:0.0185,5:0.0078,6:0.0173,7:0.0142,8:0.0078,9:0.0082,11:0.0074},
        'turnos': ['T3 DIA ANTERIOR','TURNO 1','TURNO 2','TURNO 3'],
        'turno_offset': {'T3 DIA ANTERIOR':3,'TURNO 1':4,'TURNO 2':5,'TURNO 3':6},
        'linhas_por_dia': 11,
        'row_bruto': 7, 'row_liq': 8,
    },
    'CENTRO': {
        'col_map': {'AMEX':2,'HIPER':3,'MASTER CREDITO':4,'MASTER DEBITO':5,'VISA CREDITO':6,'VISA DEBITO':7,'ELO CREDITO':8,'ELO DEBITO':9,'PIX':11},
        'taxas': {2:0.0210,3:0.0249,4:0.0185,5:0.0078,6:0.0173,7:0.0078,8:0.0142,9:0.0082,11:0.0074},
        'turnos': ['T3 DIA ANTERIOR','TURNO 1','TURNO 2','TURNO 3'],
        'turno_offset': {'T3 DIA ANTERIOR':3,'TURNO 1':4,'TURNO 2':5,'TURNO 3':6},
        'linhas_por_dia': 11,
        'row_bruto': 7, 'row_liq': 8,
    },
    'BURAQUINHO': {
        'col_map': {'AMEX':2,'HIPER':3,'MASTER CREDITO':4,'MASTER DEBITO':5,'VISA CREDITO':6,'ELO CREDITO':7,'VISA DEBITO':8,'ELO DEBITO':9,'PIX':11},
        'taxas': {2:0.0210,3:0.0249,4:0.0185,5:0.0078,6:0.0173,7:0.0142,8:0.0078,9:0.0082,11:0.0074},
        'turnos': ['T3 DIA ANTERIOR','TURNO 1','TURNO 2','TURNO 3'],
        'turno_offset': {'T3 DIA ANTERIOR':3,'TURNO 1':4,'TURNO 2':5,'TURNO 3':6},
        'linhas_por_dia': 11,
        'row_bruto': 7, 'row_liq': 8,
    },
    'ALADAH': {
        'col_map': {'AMEX':2,'HIPER':3,'MASTER CREDITO':4,'MASTER DEBITO':5,'VISA CREDITO':6,'VISA DEBITO':7,'ELO CREDITO':8,'ELO DEBITO':9,'PIX':11},
        'taxas': {2:0.0210,3:0.0249,4:0.0185,5:0.0078,6:0.0173,7:0.0078,8:0.0142,9:0.0082,11:0.0074},
        'turnos': ['TURNO 1','TURNO 2'],
        'turno_offset': {'TURNO 1':3,'TURNO 2':4},
        'linhas_por_dia': 9,
        'row_bruto': 5, 'row_liq': 6,
    },
    'PHAB': {
        'col_map': {'AMEX':2,'HIPER':3,'MASTER CREDITO':4,'MASTER DEBITO':5,'VISA CREDITO':6,'ELO CREDITO':7,'VISA DEBITO':8,'ELO DEBITO':9,'PIX':12},
        'taxas': {2:0.0210,3:0.0249,4:0.0185,5:0.0078,6:0.0173,7:0.0142,8:0.0078,9:0.0082,12:0.0074},
        'turnos': ['TURNO 1','TURNO 2'],
        'turno_offset': {'TURNO 1':2,'TURNO 2':3},
        'linhas_por_dia': 9,
        'row_bruto': 5, 'row_liq': 6,
    },
}

def to_min(s):
    h, m = map(int, s.split(':'))
    return h * 60 + m

def parse_hora(val):
    s = str(val).strip()
    m = re.match(r'^(\d{1,2}):(\d{2})', s)
    if m:
        return int(m.group(1)) * 60 + int(m.group(2))
    try:
        f = float(s)
        if f < 1:
            return round(f * 1440)
    except:
        pass
    return None

def parse_data_dia(val):
    s = str(val).strip()
    m = re.match(r'^(\d{1,2})/(\d{1,2})/(\d{4})', s)
    if m:
        return int(m.group(1))
    m = re.match(r'^(\d{4})-(\d{1,2})-(\d{1,2})', s)
    if m:
        return int(m.group(3))
    try:
        f = float(s)
        if f > 40000:
            d = date(1899, 12, 30) + timedelta(days=int(f))
            return d.day
    except:
        pass
    return None

def classifica_turno(min_val, turnos_config, t3af=None, t1i=None, t1f=None, t2i=None, t2f=None, t3i=None):
    if 'T3 DIA ANTERIOR' in turnos_config and t3af and min_val <= t3af:
        return 'T3 DIA ANTERIOR'
    if t1i and t1f and t1i <= min_val <= t1f:
        return 'TURNO 1'
    if t2i and t2f and t2i <= min_val <= t2f:
        return 'TURNO 2'
    if 'TURNO 3' in turnos_config and t3i and min_val >= t3i:
        return 'TURNO 3'
    return None

def map_bandeira(band, forma):
    b = band.lower()
    f = (forma or '').lower()
    is_d = 'débit' in f or 'debito' in f or 'débito' in f or 'debit' in f
    if 'visa' in b: return 'VISA DEBITO' if is_d else 'VISA CREDITO'
    if 'master' in b: return 'MASTER DEBITO' if is_d else 'MASTER CREDITO'
    if 'elo' in b: return 'ELO DEBITO' if is_d else 'ELO CREDITO'
    if 'amex' in b or 'american' in b: return 'AMEX'
    if 'hiper' in b: return 'HIPER'
    return None

def encontra_header(rows):
    for i, row in enumerate(rows):
        r = [str(c).strip().lower() for c in row]
        hi = next((j for j, c in enumerate(r) if c == 'hora da venda'), -1)
        if hi != -1:
            return {
                'idx': i, 'hora': hi,
                'data': next((j for j, c in enumerate(r) if c == 'data da venda'), -1),
                'valor': next((j for j, c in enumerate(r) if c == 'valor bruto'), -1),
                'band': next((j for j, c in enumerate(r) if c == 'bandeira'), -1),
                'status': next((j for j, c in enumerate(r) if c == 'status da venda'), -1),
                'forma': next((j for j, c in enumerate(r) if c == 'forma de pagamento'), -1),
            }
    return None

@app.route('/')
def index():
    return app.send_static_file('index.html')

@app.route('/preencher', methods=['POST'])
def preencher():
    try:
        planilha = request.files.get('planilha')
        cartoes = request.files.get('cartoes')
        pix = request.files.get('pix')
        dados_json = request.form.get('dados')

        if not planilha or not dados_json:
            return jsonify({'erro': 'Arquivo ou dados nao enviados'}), 400

        dados = json.loads(dados_json)
        posto = dados['posto']
        aba = dados['aba']
        dias_config = dados['dias']

        cfg = POSTO_CONFIGS.get(posto)
        if not cfg:
            return jsonify({'erro': f'Posto {posto} nao reconhecido'}), 400

        col_map = cfg['col_map']
        taxas = cfg['taxas']
        turnos = cfg['turnos']
        turno_offset = cfg['turno_offset']
        linhas_por_dia = cfg['linhas_por_dia']
        row_bruto = cfg['row_bruto']
        row_liq = cfg['row_liq']

        from collections import defaultdict
        matriz = defaultdict(lambda: defaultdict(lambda: defaultdict(float)))

        def processar_arquivo(file, tipo):
            if not file:
                return
            wb2 = openpyxl.load_workbook(file, data_only=True)
            ws2 = wb2.active
            rows = [[cell.value for cell in row] for row in ws2.iter_rows()]
            h = encontra_header(rows)
            if not h:
                return
            for row in rows[h['idx']+1:]:
                if not row[h['hora']]:
                    continue
                status = str(row[h['status']] or '').strip().lower()
                if status and status != 'aprovada':
                    continue
                min_val = parse_hora(row[h['hora']])
                if min_val is None:
                    continue
                valor = 0
                try:
                    valor = float(str(row[h['valor']]).replace(',', '.'))
                except:
                    pass
                if valor <= 0:
                    continue
                dia_tx = parse_data_dia(row[h['data']]) if h['data'] >= 0 else None
                if dia_tx is None:
                    continue
                cfg_dia = next((d for d in dias_config if d['dia'] == dia_tx), None)
                if not cfg_dia:
                    continue
                t = cfg_dia['turnos']
                turno = classifica_turno(
                    min_val, turnos,
                    t3af=to_min(t['t3af']) if t.get('t3af') else None,
                    t1i=to_min(t['t1i']), t1f=to_min(t['t1f']),
                    t2i=to_min(t['t2i']), t2f=to_min(t['t2f']),
                    t3i=to_min(t['t3i']) if t.get('t3i') else None,
                )
                if not turno:
                    continue
                coluna = 'PIX' if tipo == 'pix' else map_bandeira(str(row[h['band']] or ''), str(row[h['forma']] or ''))
                if coluna and coluna in col_map:
                    matriz[dia_tx][turno][coluna] += valor

        processar_arquivo(cartoes, 'cartoes')
        processar_arquivo(pix, 'pix')

        planilha_bytes = planilha.read()
        import io as _io
        # Ler planilha original para saber quais células tinham 0
        wb_ref = openpyxl.load_workbook(_io.BytesIO(planilha_bytes), data_only=True)
        ws_ref = wb_ref[aba] if aba in wb_ref.sheetnames else None
        wb = openpyxl.load_workbook(_io.BytesIO(planilha_bytes))
        if aba not in wb.sheetnames:
            return jsonify({'erro': f'Aba "{aba}" nao encontrada. Abas disponiveis: {", ".join(wb.sheetnames)}'}), 400
        ws = wb[aba]

        for dia_tx, turnos_dia in matriz.items():
            base = (dia_tx - 1) * linhas_por_dia
            for turno, colunas in turnos_dia.items():
                if turno not in turno_offset:
                    continue
                row = base + turno_offset[turno]
                for col_name, valor in colunas.items():
                    col_idx = col_map[col_name]
                    cell = ws.cell(row=row, column=col_idx)
                    if valor > 0:
                        cell.value = round(valor, 2)
                    else:
                        # Se a célula original tinha 0, manter 0 para preservar estrutura
                        ref_val = ws_ref.cell(row=row, column=col_idx).value if ws_ref else None
                        if ref_val == 0:
                            cell.value = 0

        for d in range(1, 32):
            base_d = (d - 1) * linhas_por_dia
            r_liq = base_d + row_liq
            r_bruto = base_d + row_bruto
            for col_idx, taxa in taxas.items():
                bruto_ref = ws.cell(row=r_bruto, column=col_idx).coordinate
                ws.cell(row=r_liq, column=col_idx).value = f'={bruto_ref}-({bruto_ref}*{taxa})'

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        dias_str = '_'.join(str(d['dia']) for d in dias_config)
        nome_arquivo = f"{posto}_{aba}_DIAS{dias_str}.xlsx"

        return send_file(output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True, download_name=nome_arquivo)

    except Exception as e:
        import traceback
        return jsonify({'erro': str(e), 'trace': traceback.format_exc()}), 500

if __name__ == '__main__':
    app.run(debug=True)
